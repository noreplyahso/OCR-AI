from __future__ import annotations

import csv
import json
from pathlib import Path

import yaml

from drb_inspection.application.contracts.context import ProductCatalogEntry


class ProductCatalogLoader:
    _FIELD_ALIASES = {
        "product_name": ("product_name", "ProductName", "Product name", "Product Name"),
        "model_path": ("model_path", "ModelPath", "Model path", "Model Path"),
        "exposure": ("exposure", "Exposure"),
        "default_number": ("default_number", "DefaultNumber", "Default number", "Default Number"),
        "threshold_accept": ("threshold_accept", "ThresholdAccept", "Threshold accept", "Threshold Accept"),
        "threshold_mns": ("threshold_mns", "ThresholdMns", "MNS threshold", "MNS Threshold", "Threshold MNS"),
    }

    def load_from_file(self, path: str | Path) -> list[ProductCatalogEntry]:
        catalog_path = Path(path)
        if not catalog_path.exists():
            raise FileNotFoundError(f"Product catalog file does not exist: {catalog_path}")

        suffix = catalog_path.suffix.lower()
        if suffix in {".yaml", ".yml"}:
            payload = yaml.safe_load(catalog_path.read_text(encoding="utf-8"))
            rows = payload.get("products", payload) if isinstance(payload, dict) else payload
            return self._rows_to_entries(rows or [])
        if suffix == ".json":
            payload = json.loads(catalog_path.read_text(encoding="utf-8"))
            rows = payload.get("products", payload) if isinstance(payload, dict) else payload
            return self._rows_to_entries(rows or [])
        if suffix == ".csv":
            with catalog_path.open("r", encoding="utf-8-sig", newline="") as handle:
                return self._rows_to_entries(list(csv.DictReader(handle)))
        if suffix in {".xlsx", ".xlsm"}:
            return self._load_excel(catalog_path)
        raise ValueError(f"Unsupported product catalog format: {catalog_path.suffix}")

    def _load_excel(self, path: Path) -> list[ProductCatalogEntry]:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError("openpyxl is required to read Excel product catalogs.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        normalized_rows: list[dict[str, object]] = []
        for values in rows[1:]:
            normalized_rows.append(
                {
                    header: value
                    for header, value in zip(headers, values)
                    if header
                }
            )
        return self._rows_to_entries(normalized_rows)

    def _rows_to_entries(self, rows: list[dict]) -> list[ProductCatalogEntry]:
        entries: list[ProductCatalogEntry] = []
        for row in rows:
            product_name = self._read_value(row, "product_name")
            if not product_name:
                continue
            entries.append(
                ProductCatalogEntry(
                    product_name=str(product_name).strip(),
                    model_path=str(self._read_value(row, "model_path") or "").strip(),
                    exposure=self._to_int(self._read_value(row, "exposure")),
                    default_number=self._to_int(self._read_value(row, "default_number")),
                    threshold_accept=self._to_float(self._read_value(row, "threshold_accept")),
                    threshold_mns=self._to_float(self._read_value(row, "threshold_mns")),
                    metadata=self._extra_metadata(row),
                )
            )
        return entries

    def _read_value(self, row: dict, logical_name: str):
        for alias in self._FIELD_ALIASES[logical_name]:
            if alias in row and row[alias] not in ("", None):
                return row[alias]
        return None

    def _extra_metadata(self, row: dict) -> dict[str, object]:
        known_columns = {
            alias
            for aliases in self._FIELD_ALIASES.values()
            for alias in aliases
        }
        return {
            str(key): value
            for key, value in row.items()
            if key not in known_columns and value not in ("", None)
        }

    @staticmethod
    def _to_int(value: object) -> int | None:
        if value in ("", None):
            return None
        return int(value)

    @staticmethod
    def _to_float(value: object) -> float | None:
        if value in ("", None):
            return None
        return float(value)
